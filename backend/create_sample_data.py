# =============================================================================
# create_sample_data.py — Generate Sample Excel Files for Testing
# =============================================================================
# This script creates realistic sample Excel files that match the structure
# shown in the screenshot, so you can test the app without needing real data.
#
# Run this script ONCE to create the sample files:
#   cd backend
#   python create_sample_data.py
#
# It creates:
#   - uploads/sample_parent.xlsx  (Russell 1000 Growth Index benchmark data)
#   - uploads/sample_child.xlsx   (Harris Corporation portfolio data)
# =============================================================================

import os
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Make sure uploads directory exists
UPLOADS_DIR = Path(__file__).parent / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)


# =============================================================================
# SAMPLE DATA — Mimics what's in the screenshot
# =============================================================================
# The "parent" is the benchmark index (Russell 1000 Growth)
PARENT_COMPANIES = [
    # sector,           company,                          ticker, avg_w, end_w, ret,    contrib
    ("Health care",     "Abbvie Inc.",                   "ABBV",  1.29,  1.35, -0.60,  -0.01),
    ("Health care",     "Natera, Inc.",                  "NTRA",  0.09,  0.10,  42.32,  0.03),
    ("Health care",     "Intuitive Surgical, Inc.",      "ISRG",  0.60,  0.66,  26.64,  0.14),
    ("Health care",     "Alnylam Pharmaceuticals, Inc.", "ALNY",  0.18,  0.17, -12.80, -0.02),
    ("Health care",     "Veeva Systems Inc.",            "VEEV",  0.10,  0.09, -25.07, -0.03),
    ("Health care",     "Zoetis Inc.",                   "ZTS",   0.15,  0.16, -13.71, -0.02),
    ("Health care",     "McKesson Corporation",          "MCK",   0.30,  0.32,   6.28,  0.02),
    ("Health care",     "Corcept Therapeutics Inc.",     "CORT",  0.02,  0.01, -58.13, -0.02),
    ("Health care",     "Idexx Laboratories, Inc.",      "IDXX",  0.18,  0.18,   5.89,  0.01),
    ("Health care",     "Doximitiy, Inc.",               "DOCS",  0.02,  0.02, -39.47, -0.01),
    ("Technology",      "Apple Inc.",                    "AAPL",  8.50,  8.70,  12.50,  0.95),
    ("Technology",      "Microsoft Corporation",         "MSFT",  9.20,  9.40,   8.30,  0.72),
    ("Technology",      "NVIDIA Corporation",            "NVDA",  7.10,  7.50,  18.40,  1.20),
    ("Technology",      "Alphabet Inc. Class A",         "GOOGL", 4.20,  4.30,   6.80,  0.27),
    ("Technology",      "Meta Platforms, Inc.",          "META",  3.80,  3.90,  14.20,  0.51),
    ("Financials",      "Visa Inc.",                     "V",     2.10,  2.20,   9.50,  0.19),
    ("Financials",      "Mastercard Incorporated",       "MA",    1.90,  2.00,   7.80,  0.14),
    ("Consumer Disc.",  "Amazon.com, Inc.",              "AMZN",  5.60,  5.80,  16.30,  0.87),
    ("Consumer Disc.",  "Tesla, Inc.",                   "TSLA",  2.30,  2.20, -18.50, -0.46),
    ("Industrials",     "Caterpillar Inc.",              "CAT",   0.80,  0.82,   4.20,  0.03),
]

# The "child" is the portfolio — some companies are there, some not,
# and the data for held positions may differ
CHILD_COMPANIES = [
    # sector,           company,                          ticker, avg_w, end_w, ret,    contrib
    ("Health care",     "Abbvie Inc.",                   "ABBV",  0.00,  0.00,  0.00,   0.00),   # Held but 0 weight (not owned)
    ("Health care",     "Natera Inc",                    "NTRA",  0.96,  1.01,  42.32,  0.33),   # Slightly different name! + weight differs
    ("Health care",     "Intuitive Surgical Inc",        "ISRG",  2.04,  2.20,  26.64,  0.46),   # Higher weight than benchmark
    ("Health care",     "Alnylam Pharmaceuticals Inc",   "ALNY",  0.00,  0.00,   0.00,  0.00),   # Not owned
    ("Health care",     "Veeva Systems",                 "VEEV",  0.00,  0.00,   0.00,  0.00),   # Not owned, shorter name
    ("Health care",     "Zoetis",                        "ZTS",   0.00,  0.00,   0.00,  0.00),   # Not owned, shorter name
    ("Health care",     "McKesson Corporation",          "MCK",   0.00,  0.00,   0.00,  0.00),   # Not owned
    ("Technology",      "Apple Inc.",                    "AAPL",  9.80, 10.10,  12.50,  1.10),   # Overweight vs benchmark
    ("Technology",      "Microsoft Corp",                "MSFT",  8.50,  8.70,   8.30,  0.66),   # Slightly different name
    ("Technology",      "Nvidia Corp",                   "NVDA",  6.20,  6.50,  18.40,  1.05),   # Different name
    ("Technology",      "Alphabet Inc Class A",          "GOOGL", 3.50,  3.60,   6.80,  0.22),   # Slightly different
    ("Financials",      "Visa Inc.",                     "V",     2.50,  2.60,   9.50,  0.22),   # Overweight
    ("Financials",      "Mastercard Incorporated",       "MA",    0.00,  0.00,   0.00,  0.00),   # Not owned
    ("Consumer Disc.",  "Amazon.com Inc",                "AMZN",  6.10,  6.30,  16.30,  0.95),   # Slight name diff + overweight
    ("Industrials",     "Caterpillar Inc",               "CAT",   1.20,  1.25,   4.20,  0.05),   # Overweight, name slightly different
    ("Industrials",     "Deere & Company",               "DE",    0.90,  0.95,   7.80,  0.07),   # Not in parent!
]


# =============================================================================
# CREATE PARENT EXCEL FILE
# =============================================================================
def create_parent_excel(filepath: str):
    """
    Creates a realistic Excel file mimicking the Performance Attribution Report format.
    
    OPENPYXL BASICS:
    - Workbook()    → Create a new Excel file in memory
    - wb.active     → Get the default sheet
    - ws['A1']      → Access cell A1
    - ws.cell(row=1, col=1).value = "text"  → Set cell value
    - ws.merge_cells('A1:D1')  → Merge cells (like Excel's "Merge & Center")
    - Font(bold=True, size=12) → Set font properties
    - wb.save(path) → Write to disk
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined"
    
    # ------------------------------------------------------------------
    # METADATA ROWS (rows 1-10) — matches the screenshot structure
    # ------------------------------------------------------------------
    ws['A1'] = ""
    ws['A2'] = "Performance Attribution - Detail"
    ws['A3'] = "Harris Corporation - Large-cap Growth (total return)"
    ws['A4'] = "Russell 1000 Growth Index (Total Return)"
    ws['A5'] = "11 Groups of TRP GICS Sector"
    ws['A6'] = "09/30/2025 - 12/31/2025"
    ws['A7'] = "Base Currency: U.S. dollar"
    ws['A8'] = ""
    ws['A9'] = ""
    
    # Style the title rows
    for row in [2, 3, 4]:
        ws[f'A{row}'].font = Font(bold=True, size=10)
    
    # ------------------------------------------------------------------
    # COLUMN HEADERS (rows 10-11) — two-row header like in the screenshot
    # ------------------------------------------------------------------
    # Portfolio header (spans columns B-G)
    ws['B10'] = "Harris Corporation - Large-cap Growth"
    ws.merge_cells('B10:G10')
    ws['B10'].font = Font(bold=True, color="FFFFFF")
    ws['B10'].fill = PatternFill("solid", fgColor="1F4E79")
    ws['B10'].alignment = Alignment(horizontal="center")
    
    # Benchmark header (spans columns H-K)
    ws['H10'] = "Russell 1000 Growth Index"
    ws.merge_cells('H10:K10')
    ws['H10'].font = Font(bold=True, color="FFFFFF")
    ws['H10'].fill = PatternFill("solid", fgColor="375623")
    ws['H10'].alignment = Alignment(horizontal="center")
    
    # Column sub-headers (row 11)
    headers = [
        "Company Name", "Exchange", "Price 12/31",
        "Average W", "Ending W", "Return (%)", "Contrib",   # Portfolio cols
        "Average W", "Ending W", "Return (%)", "Contrib",   # Benchmark cols
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=11, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    
    # ------------------------------------------------------------------
    # DATA ROWS (starting at row 13)
    # ------------------------------------------------------------------
    current_row = 13
    current_sector = None
    
    for sector, company, ticker, avg_w, end_w, ret, contrib in PARENT_COMPANIES:
        # Insert sector header row when sector changes
        if sector != current_sector:
            ws.cell(row=current_row, column=1, value=sector)
            ws.cell(row=current_row, column=1).font = Font(bold=True, italic=True)
            ws.cell(row=current_row, column=1).fill = PatternFill("solid", fgColor="FFF2CC")
            current_row += 1
            current_sector = sector
        
        # Write company data
        ws.cell(row=current_row, column=1, value=company)
        ws.cell(row=current_row, column=2, value=ticker)
        ws.cell(row=current_row, column=3, value=0)       # Price placeholder
        ws.cell(row=current_row, column=4, value=avg_w)    # Portfolio avg weight (0 = not held)
        ws.cell(row=current_row, column=5, value=end_w)    # Portfolio ending weight
        ws.cell(row=current_row, column=6, value=ret)      # Portfolio return
        ws.cell(row=current_row, column=7, value=contrib)  # Portfolio contribution
        ws.cell(row=current_row, column=8, value=avg_w)    # Benchmark avg weight
        ws.cell(row=current_row, column=9, value=end_w)    # Benchmark ending weight
        ws.cell(row=current_row, column=10, value=ret)     # Benchmark return
        ws.cell(row=current_row, column=11, value=contrib) # Benchmark contribution
        
        # Zebra striping (alternating row colors)
        fill_color = "EBF3FB" if current_row % 2 == 0 else "FFFFFF"
        for col in range(1, 12):
            ws.cell(row=current_row, column=col).fill = PatternFill("solid", fgColor=fill_color)
        
        current_row += 1
    
    # Set column widths for readability
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 8
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col].width = 12
    
    wb.save(filepath)
    print(f"✅ Created parent Excel: {filepath}")


# =============================================================================
# CREATE CHILD EXCEL FILE
# =============================================================================
def create_child_excel(filepath: str):
    """Creates the child/portfolio Excel with slightly different data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined"
    
    # Metadata rows
    ws['A2'] = "Performance Attribution - Detail"
    ws['A3'] = "Harris Corporation - Large-cap Growth (total return)"
    ws['A4'] = "Russell 1000 Growth Index (Total Return)"
    ws['A5'] = "11 Groups of TRP GICS Sector"
    ws['A6'] = "09/30/2025 - 12/31/2025"
    ws['A7'] = "Base Currency: U.S. dollar"
    
    # Headers
    ws['B10'] = "Harris Corporation - Large-cap Growth"
    ws.merge_cells('B10:G10')
    ws['B10'].font = Font(bold=True, color="FFFFFF")
    ws['B10'].fill = PatternFill("solid", fgColor="1F4E79")
    ws['B10'].alignment = Alignment(horizontal="center")
    
    headers = [
        "Company Name", "Exchange", "Price 12/31",
        "Average W", "Ending W", "Return (%)", "Contrib",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=11, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
    
    # Data rows
    current_row = 13
    current_sector = None
    
    for sector, company, ticker, avg_w, end_w, ret, contrib in CHILD_COMPANIES:
        if sector != current_sector:
            ws.cell(row=current_row, column=1, value=sector)
            ws.cell(row=current_row, column=1).font = Font(bold=True, italic=True)
            ws.cell(row=current_row, column=1).fill = PatternFill("solid", fgColor="FFF2CC")
            current_row += 1
            current_sector = sector
        
        ws.cell(row=current_row, column=1, value=company)
        ws.cell(row=current_row, column=2, value=ticker)
        ws.cell(row=current_row, column=3, value=0)
        ws.cell(row=current_row, column=4, value=avg_w)
        ws.cell(row=current_row, column=5, value=end_w)
        ws.cell(row=current_row, column=6, value=ret)
        ws.cell(row=current_row, column=7, value=contrib)
        current_row += 1
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 8
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 12
    
    wb.save(filepath)
    print(f"✅ Created child Excel: {filepath}")


# =============================================================================
# MAIN — Run when script is executed directly
# =============================================================================
if __name__ == "__main__":
    parent_path = str(UPLOADS_DIR / "sample_parent.xlsx")
    child_path  = str(UPLOADS_DIR / "sample_child.xlsx")
    
    create_parent_excel(parent_path)
    create_child_excel(child_path)
    
    print("\n🎉 Sample files created!")
    print(f"   Parent: {parent_path}")
    print(f"   Child:  {child_path}")
    print("\nNext steps:")
    print("  1. Start the backend: uvicorn main:app --reload")
    print("  2. Start the frontend: cd ../frontend && npm run dev")
    print("  3. Open http://localhost:5173 in your browser")
